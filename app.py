from flask import Flask, render_template, request, url_for, make_response, flash, redirect, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import date, datetime
import csv
import io
import os

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///expenses.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'my-secret-key'
db = SQLAlchemy(app)


class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(120), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    category = db.Column(db.String(50), nullable=False)
    date = db.Column(db.Date, nullable=False, default=date.today)


class Saving(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(120), nullable=False)
    budget = db.Column(db.Float, nullable=False, default=0.0)
    actual = db.Column(db.Float, nullable=False, default=0.0)


class Debt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(120), nullable=False)
    budget = db.Column(db.Float, nullable=False, default=0.0)
    actual = db.Column(db.Float, nullable=False, default=0.0)


class MonthlyIncome(db.Model):
    id     = db.Column(db.Integer, primary_key=True)
    year   = db.Column(db.Integer, nullable=False)
    month  = db.Column(db.Integer, nullable=False)   # 1-12
    amount = db.Column(db.Float,   nullable=False, default=0.0)
    __table_args__ = (db.UniqueConstraint('year', 'month', name='uq_year_month'),)


with app.app_context():
    db.create_all()

CATEGORIES = ['Groceries', 'Health', 'Rent', 'Insurance', 'Utilities', 'Transport', 'Others']


def parse_date_or_none(s: str):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


# ──────────────────────────────────────────────
# INDEX  –  list + filter
# ──────────────────────────────────────────────
@app.route("/")
def index():
    start_str       = (request.args.get("start")    or "").strip()
    end_str         = (request.args.get("end")      or "").strip()
    category_filter = (request.args.get("category") or "").strip()

    start_date = parse_date_or_none(start_str)
    end_date   = parse_date_or_none(end_str)

    if start_date and end_date and end_date < start_date:
        flash("End date cannot be before start date.", "error")
        start_date = end_date = None
        start_str = end_str = ""

    q = Expense.query
    if start_date:
        q = q.filter(Expense.date >= start_date)
    if end_date:
        q = q.filter(Expense.date <= end_date)
    if category_filter:
        q = q.filter(Expense.category == category_filter)

    expenses = q.order_by(Expense.date.desc(), Expense.id.desc()).all()
    total = round(sum(e.amount for e in expenses), 2)

    cat_totals: dict[str, float] = {}
    for e in expenses:
        cat_totals[e.category] = round(cat_totals.get(e.category, 0) + e.amount, 2)

    day_totals: dict[str, float] = {}
    for e in expenses:
        d = e.date.isoformat() if e.date else "Unknown"
        day_totals[d] = round(day_totals.get(d, 0) + e.amount, 2)
    day_totals = dict(sorted(day_totals.items()))

    savings = Saving.query.all()
    debts   = Debt.query.all()

    savings_budget_total = round(sum(s.budget for s in savings), 2)
    savings_actual_total = round(sum(s.actual for s in savings), 2)
    debt_budget_total    = round(sum(d.budget for d in debts), 2)
    debt_actual_total    = round(sum(d.actual for d in debts), 2)

    # Income for selected/current month
    import calendar as _cal
    today_dt  = date.today()
    inc_year  = int(request.args.get("inc_year",  today_dt.year))
    inc_month = int(request.args.get("inc_month", today_dt.month))
    inc_row   = MonthlyIncome.query.filter_by(year=inc_year, month=inc_month).first()
    current_income = inc_row.amount if inc_row else 0.0

    # Income vs Debt chart – last 6 months
    income_chart = []
    debt_chart   = []
    labels_chart = []
    for i in range(5, -1, -1):
        offset = today_dt.month - 1 - i
        y = today_dt.year + offset // 12
        m = offset % 12 + 1
        row = MonthlyIncome.query.filter_by(year=y, month=m).first()
        income_chart.append(round(row.amount if row else 0.0, 2))
        debt_chart.append(debt_actual_total)
        labels_chart.append(f"{_cal.month_abbr[m]} {y}")

    return render_template(
        "index.html",
        categories=CATEGORIES,
        today=date.today().isoformat(),
        expenses=expenses,
        total=total,
        start_str=start_str,
        end_str=end_str,
        category_filter=category_filter,
        cat_totals=cat_totals,
        day_totals=day_totals,
        savings=savings,
        debts=debts,
        savings_budget_total=savings_budget_total,
        savings_actual_total=savings_actual_total,
        debt_budget_total=debt_budget_total,
        debt_actual_total=debt_actual_total,
        current_income=current_income,
        inc_year=inc_year,
        inc_month=inc_month,
        income_chart_labels=labels_chart,
        income_chart_data=income_chart,
        debt_chart_data=debt_chart,
    )


# ──────────────────────────────────────────────
# ADD EXPENSE
# ──────────────────────────────────────────────
@app.route("/add", methods=["POST"])
def add():
    description = (request.form.get("description") or "").strip()
    amount_str  = (request.form.get("amount")      or "").strip()
    category    = (request.form.get("category")    or "").strip()
    date_str    = (request.form.get("date")        or "").strip()

    if not description or not amount_str or not category:
        flash("Please fill in description, amount, and category.", "error")
        return redirect(url_for("index"))

    try:
        amount = float(amount_str)
        if amount <= 0:
            raise ValueError
    except ValueError:
        flash("Amount must be a positive number.", "error")
        return redirect(url_for("index"))

    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else date.today()
    except ValueError:
        d = date.today()

    db.session.add(Expense(description=description, amount=amount, category=category, date=d))
    db.session.commit()
    flash("Expense added successfully.", "success")
    return redirect(url_for("index"))


# ──────────────────────────────────────────────
# DELETE EXPENSE
# ──────────────────────────────────────────────
@app.route("/delete/<int:expense_id>", methods=["POST"])
def delete(expense_id):
    e = Expense.query.get_or_404(expense_id)
    db.session.delete(e)
    db.session.commit()
    flash("Expense deleted.", "success")
    return redirect(url_for("index"))


# ──────────────────────────────────────────────
# EDIT EXPENSE
# ──────────────────────────────────────────────
@app.route("/edit/<int:expense_id>", methods=["GET"])
def edit_get(expense_id):
    e = Expense.query.get_or_404(expense_id)
    return render_template("edit.html", expense=e, categories=CATEGORIES, today=date.today().isoformat())


@app.route("/edit/<int:expense_id>", methods=["POST"])
def edit_post(expense_id):
    e = Expense.query.get_or_404(expense_id)

    description = (request.form.get("description") or "").strip()
    amount_str  = (request.form.get("amount")      or "").strip()
    category    = (request.form.get("category")    or "").strip()
    date_str    = (request.form.get("date")        or "").strip()

    errors = []
    if not description:
        errors.append("Description is required.")
    if not amount_str:
        errors.append("Amount is required.")
    if not category:
        errors.append("Category is required.")

    amount = None
    if amount_str:
        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError
        except ValueError:
            errors.append("Amount must be a positive number.")

    if errors:
        for err in errors:
            flash(err, "error")
        return render_template("edit.html", expense=e, categories=CATEGORIES, today=date.today().isoformat())

    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else e.date
    except ValueError:
        d = e.date

    e.description = description
    e.amount      = amount
    e.category    = category
    e.date        = d
    db.session.commit()
    flash("Expense updated successfully.", "success")
    return redirect(url_for("index"))


# ──────────────────────────────────────────────
# SAVINGS CRUD
# ──────────────────────────────────────────────
@app.route("/savings/add", methods=["POST"])
def add_saving():
    description = (request.form.get("description") or "").strip()
    budget_str  = (request.form.get("budget")      or "0").strip()
    actual_str  = (request.form.get("actual")      or "0").strip()

    if not description:
        flash("Saving description is required.", "error")
        return redirect(url_for("index") + "#savings")

    try:
        budget = float(budget_str) if budget_str else 0.0
        actual = float(actual_str) if actual_str else 0.0
    except ValueError:
        flash("Budget and actual must be numbers.", "error")
        return redirect(url_for("index") + "#savings")

    db.session.add(Saving(description=description, budget=budget, actual=actual))
    db.session.commit()
    flash("Saving added.", "success")
    return redirect(url_for("index") + "#savings")


@app.route("/savings/delete/<int:saving_id>", methods=["POST"])
def delete_saving(saving_id):
    s = Saving.query.get_or_404(saving_id)
    db.session.delete(s)
    db.session.commit()
    flash("Saving deleted.", "success")
    return redirect(url_for("index") + "#savings")


@app.route("/savings/edit/<int:saving_id>", methods=["POST"])
def edit_saving(saving_id):
    s = Saving.query.get_or_404(saving_id)
    s.description = (request.form.get("description") or s.description).strip()
    try:
        s.budget = float(request.form.get("budget") or 0)
        s.actual = float(request.form.get("actual") or 0)
    except ValueError:
        flash("Invalid number.", "error")
        return redirect(url_for("index") + "#savings")
    db.session.commit()
    flash("Saving updated.", "success")
    return redirect(url_for("index") + "#savings")


# ──────────────────────────────────────────────
# DEBT CRUD
# ──────────────────────────────────────────────
@app.route("/debt/add", methods=["POST"])
def add_debt():
    description = (request.form.get("description") or "").strip()
    budget_str  = (request.form.get("budget")      or "0").strip()
    actual_str  = (request.form.get("actual")      or "0").strip()

    if not description:
        flash("Debt description is required.", "error")
        return redirect(url_for("index") + "#debt")

    try:
        budget = float(budget_str) if budget_str else 0.0
        actual = float(actual_str) if actual_str else 0.0
    except ValueError:
        flash("Budget and actual must be numbers.", "error")
        return redirect(url_for("index") + "#debt")

    db.session.add(Debt(description=description, budget=budget, actual=actual))
    db.session.commit()
    flash("Debt added.", "success")
    return redirect(url_for("index") + "#debt")


@app.route("/debt/delete/<int:debt_id>", methods=["POST"])
def delete_debt(debt_id):
    d = Debt.query.get_or_404(debt_id)
    db.session.delete(d)
    db.session.commit()
    flash("Debt deleted.", "success")
    return redirect(url_for("index") + "#debt")


@app.route("/debt/edit/<int:debt_id>", methods=["POST"])
def edit_debt(debt_id):
    d = Debt.query.get_or_404(debt_id)
    d.description = (request.form.get("description") or d.description).strip()
    try:
        d.budget = float(request.form.get("budget") or 0)
        d.actual = float(request.form.get("actual") or 0)
    except ValueError:
        flash("Invalid number.", "error")
        return redirect(url_for("index") + "#debt")
    db.session.commit()
    flash("Debt updated.", "success")
    return redirect(url_for("index") + "#debt")


# ──────────────────────────────────────────────
# MONTHLY INCOME  –  upsert by year/month
# ──────────────────────────────────────────────
@app.route("/income/set", methods=["POST"])
def set_income():
    import calendar as _cal
    year_str   = (request.form.get("inc_year")   or "").strip()
    month_str  = (request.form.get("inc_month")  or "").strip()
    amount_str = (request.form.get("inc_amount") or "").strip()

    today_dt = date.today()
    try:
        year  = int(year_str)  if year_str  else today_dt.year
        month = int(month_str) if month_str else today_dt.month
        if not (1 <= month <= 12):
            raise ValueError
    except ValueError:
        flash("Invalid month/year.", "error")
        return redirect(url_for("index"))

    try:
        amount = float(amount_str)
        if amount < 0:
            raise ValueError
    except ValueError:
        flash("Income must be a non-negative number.", "error")
        return redirect(url_for("index"))

    row = MonthlyIncome.query.filter_by(year=year, month=month).first()
    if row:
        row.amount = amount
    else:
        db.session.add(MonthlyIncome(year=year, month=month, amount=amount))
    db.session.commit()
    flash(f"Income for {_cal.month_name[month]} {year} saved.", "success")
    return redirect(url_for("index", inc_year=year, inc_month=month))


# ──────────────────────────────────────────────
# EXPORT CSV  –  respects current filters
# ──────────────────────────────────────────────
@app.route("/export")
def export_csv():
    start_str       = (request.args.get("start")    or "").strip()
    end_str         = (request.args.get("end")      or "").strip()
    category_filter = (request.args.get("category") or "").strip()

    start_date = parse_date_or_none(start_str)
    end_date   = parse_date_or_none(end_str)

    q = Expense.query
    if start_date:
        q = q.filter(Expense.date >= start_date)
    if end_date:
        q = q.filter(Expense.date <= end_date)
    if category_filter:
        q = q.filter(Expense.category == category_filter)

    expenses = q.order_by(Expense.date.desc(), Expense.id.desc()).all()

    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(["Date", "Description", "Category", "Amount"])
    for e in expenses:
        writer.writerow([
            e.date.isoformat() if e.date else "",
            e.description,
            e.category,
            f"{e.amount:.2f}",
        ])

    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=expenses.csv"
    output.headers["Content-Type"] = "text/csv"
    return output


# ──────────────────────────────────────────────
# EXPORT EXCEL  –  Monthly Budget Sheet layout
# ──────────────────────────────────────────────
@app.route("/export/excel")
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter

    start_str       = (request.args.get("start")    or "").strip()
    end_str         = (request.args.get("end")      or "").strip()
    category_filter = (request.args.get("category") or "").strip()

    start_date = parse_date_or_none(start_str)
    end_date   = parse_date_or_none(end_str)

    q = Expense.query
    if start_date:
        q = q.filter(Expense.date >= start_date)
    if end_date:
        q = q.filter(Expense.date <= end_date)
    if category_filter:
        q = q.filter(Expense.category == category_filter)
    expenses = q.order_by(Expense.date.desc(), Expense.id.desc()).all()

    savings = Saving.query.all()
    debts   = Debt.query.all()

    # Pull income for the requested month
    import calendar as _cal
    today_dt  = date.today()
    try:
        inc_year  = int(request.args.get("inc_year",  today_dt.year))
        inc_month = int(request.args.get("inc_month", today_dt.month))
    except ValueError:
        inc_year, inc_month = today_dt.year, today_dt.month
    inc_row = MonthlyIncome.query.filter_by(year=inc_year, month=inc_month).first()
    income_amount = inc_row.amount if inc_row else 0.0
    month_label   = f"{_cal.month_name[inc_month]} {inc_year}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Budget"

    # ── Styles ──────────────────────────────────
    CREAM   = "F2F0D8"
    PALE    = "EAE8C8"
    WHITE   = "FFFFFF"
    DARK    = "1A1A1A"

    thin = Side(style="thin", color="AAAAAA")
    med  = Side(style="medium", color="888888")
    def border(left=thin, right=thin, top=thin, bottom=thin):
        return Border(left=left, right=right, top=top, bottom=bottom)

    def hdr_style(ws, cell_ref, text, bold=True, size=11, bg=CREAM, align="center"):
        c = ws[cell_ref]
        c.value = text
        c.font  = Font(bold=bold, size=size, name="Arial")
        c.fill  = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border()

    def data_cell(ws, cell_ref, value="", bold=False, align="left", fmt=None, bg=WHITE):
        c = ws[cell_ref]
        c.value = value
        c.font  = Font(bold=bold, size=10, name="Arial")
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border()
        if fmt:
            c.number_format = fmt
        if bg != WHITE:
            c.fill = PatternFill("solid", fgColor=bg)

    MONEY = '$#,##0.00'

    # ── Column widths ────────────────────────────
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    # ── Row heights ──────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6

    # ── Title ────────────────────────────────────
    ws.merge_cells("B1:H1")
    title_cell = ws["B1"]
    title_cell.value = "MONTHLY BUDGET SHEET"
    title_cell.font  = Font(bold=True, size=16, name="Arial")
    title_cell.fill  = PatternFill("solid", fgColor=CREAM)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border(left=med, right=med, top=med, bottom=med)

    # ─────────────────────────────────────────────
    # LEFT COLUMN: BILLS  (B:D, rows 3..N)
    # ─────────────────────────────────────────────
    BILLS_HDR = 3
    ws.merge_cells(f"B{BILLS_HDR}:D{BILLS_HDR}")
    hdr_style(ws, f"B{BILLS_HDR}", "BILLS", size=12, bg=PALE)
    ws[f"C{BILLS_HDR}"].fill = PatternFill("solid", fgColor=PALE)
    ws[f"D{BILLS_HDR}"].fill = PatternFill("solid", fgColor=PALE)

    r = BILLS_HDR + 1
    for lbl, col in [("BILLS", "B"), ("BUDGET", "C"), ("ACTUAL", "D")]:
        data_cell(ws, f"{col}{r}", lbl, bold=True, align="center", bg=CREAM)
    r += 1

    bill_start = r
    for e in expenses:
        data_cell(ws, f"B{r}", e.description)
        data_cell(ws, f"C{r}", e.amount, fmt=MONEY)
        data_cell(ws, f"D{r}", e.amount, fmt=MONEY)
        r += 1

    # Pad to at least 15 rows
    while r < bill_start + 15:
        data_cell(ws, f"B{r}", ""); data_cell(ws, f"C{r}", ""); data_cell(ws, f"D{r}", "")
        r += 1

    bill_end = r - 1
    # Total row
    data_cell(ws, f"B{r}", "TOTAL", bold=True, align="center", bg=CREAM)
    c = ws[f"C{r}"]
    c.value = f"=SUM(C{bill_start}:C{bill_end})"
    c.font  = Font(bold=True, name="Arial"); c.number_format = MONEY
    c.fill  = PatternFill("solid", fgColor=CREAM); c.border = border()
    c.alignment = Alignment(horizontal="right")
    c2 = ws[f"D{r}"]
    c2.value = f"=SUM(D{bill_start}:D{bill_end})"
    c2.font  = Font(bold=True, name="Arial"); c2.number_format = MONEY
    c2.fill  = PatternFill("solid", fgColor=CREAM); c2.border = border()
    c2.alignment = Alignment(horizontal="right")
    bills_total_row = r
    r += 1

    # ─────────────────────────────────────────────
    # RIGHT COLUMN: INCOME  (F:H, rows 3..8)
    # ─────────────────────────────────────────────
    INC_R = 3
    ws.merge_cells(f"F{INC_R}:H{INC_R}")
    hdr_style(ws, f"F{INC_R}", "INCOME", size=12, bg=PALE)
    ws[f"G{INC_R}"].fill = PatternFill("solid", fgColor=PALE)
    ws[f"H{INC_R}"].fill = PatternFill("solid", fgColor=PALE)

    data_cell(ws, f"F{INC_R+1}", "MONTH", bold=True, align="center", bg=CREAM)
    data_cell(ws, f"G{INC_R+1}", month_label or "", align="center")
    ws.merge_cells(f"G{INC_R+1}:H{INC_R+1}")

    for i in range(2, 5):
        data_cell(ws, f"F{INC_R+i}", ""); data_cell(ws, f"G{INC_R+i}", "")
        ws.merge_cells(f"G{INC_R+i}:H{INC_R+i}")

    inc_total_row = INC_R + 5
    data_cell(ws, f"F{inc_total_row}", "TOTAL INCOME", bold=True, bg=CREAM)
    ti = ws[f"G{inc_total_row}"]
    ti.value = income_amount; ti.number_format = MONEY; ti.border = border()
    ti.font = Font(bold=True, name="Arial"); ti.alignment = Alignment(horizontal="right")
    ws.merge_cells(f"G{inc_total_row}:H{inc_total_row}")

    # ─────────────────────────────────────────────
    # RIGHT: SAVINGS  (F:H)
    # ─────────────────────────────────────────────
    SAV_HDR = inc_total_row + 2
    ws.merge_cells(f"F{SAV_HDR}:H{SAV_HDR}")
    hdr_style(ws, f"F{SAV_HDR}", "SAVINGS", size=12, bg=PALE)
    ws[f"G{SAV_HDR}"].fill = PatternFill("solid", fgColor=PALE)
    ws[f"H{SAV_HDR}"].fill = PatternFill("solid", fgColor=PALE)

    sr = SAV_HDR + 1
    for lbl, col in [("SAVINGS", "F"), ("BUDGET", "G"), ("ACTUAL", "H")]:
        data_cell(ws, f"{col}{sr}", lbl, bold=True, align="center", bg=CREAM)
    sr += 1

    sav_start = sr
    for s in savings:
        data_cell(ws, f"F{sr}", s.description)
        data_cell(ws, f"G{sr}", s.budget, fmt=MONEY)
        data_cell(ws, f"H{sr}", s.actual, fmt=MONEY)
        sr += 1

    while sr < sav_start + 8:
        data_cell(ws, f"F{sr}", ""); data_cell(ws, f"G{sr}", ""); data_cell(ws, f"H{sr}", "")
        sr += 1

    sav_end = sr - 1
    data_cell(ws, f"F{sr}", "TOTAL", bold=True, align="center", bg=CREAM)
    for col, start, end in [("G", sav_start, sav_end), ("H", sav_start, sav_end)]:
        tc = ws[f"{col}{sr}"]
        tc.value = f"=SUM({col}{start}:{col}{end})"
        tc.font  = Font(bold=True, name="Arial"); tc.number_format = MONEY
        tc.fill  = PatternFill("solid", fgColor=CREAM); tc.border = border()
        tc.alignment = Alignment(horizontal="right")
    sav_total_row = sr
    sr += 1

    # ─────────────────────────────────────────────
    # RIGHT: DEBT  (F:H)
    # ─────────────────────────────────────────────
    DEBT_HDR = sr + 1
    ws.merge_cells(f"F{DEBT_HDR}:H{DEBT_HDR}")
    hdr_style(ws, f"F{DEBT_HDR}", "DEBT", size=12, bg=PALE)
    ws[f"G{DEBT_HDR}"].fill = PatternFill("solid", fgColor=PALE)
    ws[f"H{DEBT_HDR}"].fill = PatternFill("solid", fgColor=PALE)

    dr = DEBT_HDR + 1
    for lbl, col in [("DEBT", "F"), ("BUDGET", "G"), ("ACTUAL", "H")]:
        data_cell(ws, f"{col}{dr}", lbl, bold=True, align="center", bg=CREAM)
    dr += 1

    debt_start = dr
    for d in debts:
        data_cell(ws, f"F{dr}", d.description)
        data_cell(ws, f"G{dr}", d.budget, fmt=MONEY)
        data_cell(ws, f"H{dr}", d.actual, fmt=MONEY)
        dr += 1

    while dr < debt_start + 8:
        data_cell(ws, f"F{dr}", ""); data_cell(ws, f"G{dr}", ""); data_cell(ws, f"H{dr}", "")
        dr += 1

    debt_end = dr - 1
    data_cell(ws, f"F{dr}", "TOTAL", bold=True, align="center", bg=CREAM)
    for col in ["G", "H"]:
        tc = ws[f"{col}{dr}"]
        tc.value = f"=SUM({col}{debt_start}:{col}{debt_end})"
        tc.font  = Font(bold=True, name="Arial"); tc.number_format = MONEY
        tc.fill  = PatternFill("solid", fgColor=CREAM); tc.border = border()
        tc.alignment = Alignment(horizontal="right")
    debt_total_row = dr
    dr += 1

    # ─────────────────────────────────────────────
    # RIGHT: NOTES
    # ─────────────────────────────────────────────
    NOTES_HDR = dr + 1
    ws.merge_cells(f"F{NOTES_HDR}:H{NOTES_HDR}")
    hdr_style(ws, f"F{NOTES_HDR}", "NOTES", size=12, bg=PALE)
    ws[f"G{NOTES_HDR}"].fill = PatternFill("solid", fgColor=PALE)
    ws[f"H{NOTES_HDR}"].fill = PatternFill("solid", fgColor=PALE)
    for nr in range(NOTES_HDR + 1, NOTES_HDR + 5):
        ws.merge_cells(f"F{nr}:H{nr}")
        data_cell(ws, f"F{nr}", "")

    # ─────────────────────────────────────────────
    # MONTHLY INCOME LEFT  (bottom right)
    # ─────────────────────────────────────────────
    MIL_HDR = NOTES_HDR + 6
    ws.merge_cells(f"F{MIL_HDR}:H{MIL_HDR}")
    hdr_style(ws, f"F{MIL_HDR}", "MONTHLY INCOME LEFT", size=11, bg=PALE)
    ws[f"G{MIL_HDR}"].fill = PatternFill("solid", fgColor=PALE)
    ws[f"H{MIL_HDR}"].fill = PatternFill("solid", fgColor=PALE)

    mr = MIL_HDR + 1
    data_cell(ws, f"F{mr}", "")
    data_cell(ws, f"G{mr}", "BUDGET", bold=True, align="center", bg=CREAM)
    data_cell(ws, f"H{mr}", "ACTUAL", bold=True, align="center", bg=CREAM)
    mr += 1
    data_cell(ws, f"F{mr}", "")
    for col, bills_col, sav_col, debt_col in [
        ("G", "C", "G", "G"),
        ("H", "D", "H", "H"),
    ]:
        tc = ws[f"{col}{mr}"]
        tc.value = (f"=G{inc_total_row}-C{bills_total_row}-G{sav_total_row}-G{debt_total_row}"
                    if col == "G" else
                    f"=G{inc_total_row}-D{bills_total_row}-H{sav_total_row}-H{debt_total_row}")
        tc.font  = Font(bold=True, name="Arial"); tc.number_format = MONEY
        tc.fill  = PatternFill("solid", fgColor=CREAM); tc.border = border()
        tc.alignment = Alignment(horizontal="right")

    # ── Stream to response ───────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = make_response(buf.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=monthly_budget.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


if __name__ == "__main__":
    app.run(debug=True, port=4848)  